import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from django.http import FileResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Sum, Avg, Q
from django.utils import timezone
from datetime import date, datetime, timedelta
from decimal import Decimal
from django.urls import reverse
from django.contrib.auth.views import LoginView
from django.views.decorators.http import require_POST
import logging
import json
from dateutil.relativedelta import relativedelta
from django.db.models.functions import TruncMonth
from django.db.models.functions import ExtractMonth, ExtractYear
from django.contrib.admin.models import LogEntry
from django.core.paginator import Paginator
from django.utils.safestring import mark_safe
from django.utils.timezone import now
from django.shortcuts import redirect
from django.contrib import messages
from django.db import IntegrityError
from .forms import AdherantForm
from .models import CompteAdherent
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from .models import LogConnexionAdherent
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import date
from .models import Adherant
from .forms import DemandePretForm
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.models import Group
from django.shortcuts import render
from collections import defaultdict
from .models import CotisationMensuelle, Adherant
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Information
from .forms import InformationForm
from django.shortcuts import render
from django.core.mail import EmailMessage
from django.conf import settings
from .forms import ContactForm
from .models import Contact

from django.shortcuts import render
from django.core.mail import EmailMessage
from django.conf import settings

from .forms import ContactForm
from .models import Contact

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from datetime import date
from .models import Adherant, Pret, CotisationMensuelle
from django.shortcuts import get_object_or_404



from .models import (
    Adherant,
    AncienAdherent,
    CotisationMensuelle,
    Cotisation,
    Pret,
    DemandePret,
    Versement,
    PaiementPret,
    Decaissement,
    ParametreCaisse,
    GalerieImage,
    Contact,
    Information
)

from .forms import (
    AdherantForm,
    DemandePretForm,
    CotisationMensuelleForm,
    RechercheAdherantForm,
    ContactForm,
    InformationForm,
    VerifierEtatForm,
    DecaissementForm
)

# ============================================================
# DÉCORATEUR RÔLES BUREAU
# ============================================================
ROLES_BUREAU = ['President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte']

def role_required(*roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login_admin')
            if request.user.is_superuser or request.user.is_staff:
                return view_func(request, *args, **kwargs)
            if request.user.groups.filter(name__in=roles).exists():
                return view_func(request, *args, **kwargs)
            messages.error(request, "⛔ Vous n'avez pas les droits pour accéder à cette page.")
            return redirect('dashboard')
        return wrapper
    return decorator

# ✅ Gérer adhérents — Président uniquement
@role_required('President')
def liste_adherants(request):
    adherants = Adherant.objects.filter(statut='actif').order_by('nom')
    return render(request, "liste_adherants.html", {"liste_adherants": adherants})

@role_required('President')
def ajouter_adherant(request):

    if request.method == 'POST':
        form = AdherantForm(request.POST)

        if form.is_valid():
            try:
                form.save()
                messages.success(request, "✅ Adhérent ajouté avec succès.")
            except IntegrityError:
                messages.error(request, "❌ ADHÉRENT DÉJÀ EXISTANT.")
        else:
            messages.error(request, "❌ Corrigez les erreurs du formulaire.")

    # 👉 IMPORTANT : rediriger vers next_url ou liste_adherants
    next_url = request.POST.get('next', 'liste_adherants')
    return redirect(next_url)

@role_required('President')
def modifier_adherant(request, pk):
    adherant = get_object_or_404(Adherant, pk=pk)

    if request.method == 'POST':
        adherant.nom            = request.POST.get('nom')
        adherant.prenom         = request.POST.get('prenom')
        adherant.email          = request.POST.get('email')
        adherant.telephone      = request.POST.get('telephone')
        adherant.adresse_rue    = request.POST.get('adresse_rue')
        adherant.departement    = request.POST.get('departement')
        adherant.sous_departement = request.POST.get('sous_departement')
        adherant.save()

        return redirect('liste_adherants')  # ← retour à la liste

    # GET → ne devrait pas arriver (modal only)
    return redirect('liste_adherants')

@role_required('President')    
def supprimer_adherant(request, pk):
    adherant = Adherant.objects.get(pk=pk)
    adherant.delete()
    return redirect('liste_adherants')

def home(request):
    return render(request, 'app_social/home.html')


# Vue index
def index(request):
    return render(request, 'app_social/index.html') # Page index

def le_president(request):
    return render(request, 'app_social/le_president.html')

def accueil(request):
    return render(request, 'app_social/accueil.html')


# PAGE INFORMATIONS
def informations(request):

    informations = Information.objects.all().order_by('-date_pub')

    derniere_information = None
    autres_informations = []

    if informations.exists():
        derniere_information = informations.first()
        autres_informations = informations[1:4]

    context = {
        'derniere_information': derniere_information,
        'autres_informations': autres_informations,
    }

    return render(
        request,
        'app_social/informations.html',
        context
    )


# AJOUTER UNE INFORMATION
@login_required
def ajouter_information(request):

    user = request.user

    # Vérification des permissions
    if not (
        user.is_superuser
        or user.groups.filter(name='President').exists()
    ):
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = InformationForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.success(request, "Information ajoutée avec succès.")
            return redirect('information')

    else:
        form = InformationForm()

    return render(request, 'app_social/ajouter_information.html', {
        'form': form
    })


# MODIFIER UNE INFORMATION
@login_required
def modifier_information(request, pk):

    user = request.user

    if not (
        user.is_superuser
        or user.groups.filter(name='President').exists()
    ):
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')

    info = get_object_or_404(Information, pk=pk)

    if request.method == 'POST':
        form = InformationForm(
            request.POST,
            request.FILES,
            instance=info
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Publication modifiée.")
            return redirect('information')

    else:
        form = InformationForm(instance=info)

    return render(request, 'app_social/ajouter_information.html', {
        'form': form
    })


# SUPPRIMER UNE INFORMATION
@login_required
def supprimer_information(request, pk):

    user = request.user

    if not (
        user.is_superuser
        or user.groups.filter(name='President').exists()
    ):
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')

    info = get_object_or_404(Information, pk=pk)

    info.delete()

    messages.success(request, "Publication supprimée.")

    return redirect('information')

def galerie(request):
    return render(request, 'galerie.html')

def information_view(request):
    informations = Information.objects.order_by('-date_pub')[:3]   
    return render(request, 'app_social/informations.html', {'informations': informations})

def modifier_information(request, id):

    info = get_object_or_404(Information, id=id)

    form = InformationForm(request.POST or None,
                           request.FILES or None,
                           instance=info)

    if form.is_valid():
        form.save()
        return redirect('information')

    return render(request,
                  'app_social/ajouter_information.html',
                  {'form': form})

def supprimer_information(request, id):

    info = get_object_or_404(Information, id=id)

    info.delete()

    return redirect('information')


def export_adherants_xlsx(request):
    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adhérents"

    # Ajouter les en-têtes
    ws.append(["Nom", "Prénom", "Email", "Téléphone"])

    # Récupérer les adhérents depuis la base de données
    adherants = Adherant.objects.all()
    for adherant in adherants:
        ws.append([adherant.nom, adherant.prenom, adherant.email, adherant.telephone])

    # Générer la réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=adherants.xlsx'
    wb.save(response)
    return response


def generate_excel(request):
    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adhérents"

    # Ajouter les en-têtes des colonnes
    ws.append(["Nom", "Prénom", "Email", "Téléphone"])

    # Récupérer les adhérents depuis la base de données
    adherants = Adherant.objects.all()
    for adherant in adherants:
        ws.append([adherant.nom, adherant.prenom, adherant.email, adherant.telephone])

    # Générer la réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=adherants.xlsx'
    wb.save(response)
    return response


def cotisation_mensuelle_view(request):
    adherant = None
    form_recherche = RechercheAdherantForm(request.POST or None)
    form_cotisation = CotisationMensuelleForm()

    # Recherche d'adhérent
    if request.method == 'POST' and 'rechercher' in request.POST:
        if form_recherche.is_valid():
            query = form_recherche.cleaned_data['query']
            adherant = Adherant.objects.filter(
                models.Q(nom__icontains=query) |
                models.Q(telephone__icontains=query)
            ).first()
            if not adherant:
                messages.error(request, "Aucun adhérent trouvé.")
    
    # Ajout cotisation
    elif request.method == 'POST' and 'ajouter_cotisation' in request.POST:
        form_cotisation = CotisationMensuelleForm(request.POST)
        if form_cotisation.is_valid():
            query = request.POST.get('query')
            adherant = Adherant.objects.filter(
                models.Q(nom__icontains=query) |
                models.Q(telephone__icontains=query)
            ).first()

            if not adherant:
                messages.error(request, "Aucun adhérent sélectionné pour la cotisation.")
                return redirect('cotisation_mensuelle')

            mois = form_cotisation.cleaned_data['mois']
            annee = form_cotisation.cleaned_data['annee']
            if CotisationMensuelle.objects.filter(adherant=adherant, mois=mois, annee=annee).exists():
                messages.warning(request, "Cet adhérent a déjà cotisé pour ce mois.")
                return redirect('cotisation_mensuelle')

            cotisation = form_cotisation.save(commit=False)
            cotisation.adherant = adherant
            cotisation.montant = 5000
            cotisation.save()

            adherant.date_dernier_cotisation = date.today()
            adherant.save()

            messages.success(request, "Cotisation enregistrée avec succès.")
            return redirect('cotisation_mensuelle')

    return render(request, 'cotisation_mensuelle.html', {
        'form_recherche': form_recherche,
        'form_cotisation': form_cotisation,
        'adherant': adherant
    })


@role_required('President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte')
def liste_cotisations_annuelles(request):
    # 🔹 Récupérer toutes les cotisations
    cotisations = CotisationMensuelle.objects.select_related('adherant').all()

    # 🔹 Stockage des mois uniques par adhérent
    data_dict = defaultdict(set)
    for c in cotisations:
        if c.mois:
            data_dict[c.adherant.id].add(c.mois)

    # 🔹 Construction des données finales
    data = []
    mois_actuel = timezone.now().month

    for adherant in Adherant.objects.all():
        mois_payes = len(data_dict[adherant.id])
        pourcentage = int((mois_payes / 12) * 100) if mois_payes > 0 else 0
        a_jour = mois_payes >= mois_actuel

        data.append({
            "nom": adherant.nom,
            "prenom": adherant.prenom,
            "departement": adherant.departement,
            "paye": mois_payes,
            "pourcentage": pourcentage,
            "a_jour": a_jour
        })

    # 🔥 TRI : NON À JOUR EN HAUT
    data = sorted(data, key=lambda x: x["a_jour"])

    return render(request, "app_social/liste_cotisations_annuelles.html", {
        "data": data,
        "adherants": Adherant.objects.all()
    })    

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import CotisationMensuelle, Adherant
from django.db.models import Case, When, IntegerField
from datetime import datetime


@login_required(login_url='login_adherent')
def verifier_etat(request):

    try:
        adherant = Adherant.objects.get(telephone=request.user)
    except Adherant.DoesNotExist:
        return redirect('login_adherent')

    # ordre réel des mois
    mois_ordre = Case(
        When(mois="Janvier", then=1),
        When(mois="Février", then=2),
        When(mois="Mars", then=3),
        When(mois="Avril", then=4),
        When(mois="Mai", then=5),
        When(mois="Juin", then=6),
        When(mois="Juillet", then=7),
        When(mois="Août", then=8),
        When(mois="Septembre", then=9),
        When(mois="Octobre", then=10),
        When(mois="Novembre", then=11),
        When(mois="Décembre", then=12),
        output_field=IntegerField()
    )

    cotisations = (
        CotisationMensuelle.objects
        .filter(adherant=adherant)
        .annotate(mois_num=mois_ordre)
        .order_by('annee', 'mois_num')
    )

    derniere = cotisations.first()

    result = None
    message = ""

    if derniere:

        # mois actuel
        mois_actuel = datetime.now().month
        annee_actuelle = datetime.now().year

        # cotisation à jour ?
        if (
            derniere.annee == annee_actuelle and
            derniere.mois_num == mois_actuel
        ):
            message = "VOUS ÊTES À JOUR"
        else:
            message = "VEUILLEZ RÉGULARISER"

        result = {
            "mois": derniere.mois,
            "annee": derniere.annee
        }

    else:
        message = "Aucune cotisation trouvée"

    return render(request, "app_social/verifier_etat.html", {
        "cotisations": cotisations,
        "result": result,
        "message": message,
        "adherant": adherant
    })
    
def galerie(request):
    images = GalerieImage.objects.all()  # Récupérer toutes les images
    print(images)  # Vérifier si les images sont bien chargées
    return render(request, 'app_social/galerie.html', {'galerie_images': images})


def nous_ecrire(request):

    success = False

    if request.method == 'POST':

        form = ContactForm(request.POST)

        if form.is_valid():

            # =========================
            # Récupération des données
            # =========================
            nom = form.cleaned_data['nom']
            email_utilisateur = form.cleaned_data['email']
            objet = form.cleaned_data['objet']
            message = form.cleaned_data['message']

            # =========================
            # Sauvegarde en base
            # =========================
            Contact.objects.create(
                nom=nom,
                email=email_utilisateur,
                objet=objet,
                message=message
            )

            # =========================
            # Contenu du mail reçu
            # =========================
            contenu = f"""
Nouveau message depuis le site Caisse Sociale DIRTECH

Nom : {nom}

Email : {email_utilisateur}

Objet : {objet}

Message :
{message}
"""

            # =========================
            # Email envoyé à l'administration
            # =========================
            email_admin = EmailMessage(
                subject=f"Caisse Sociale DIRTECH - {objet}",
                body=contenu,

                # Nom affiché comme expéditeur
                from_email=f"{nom} <{settings.EMAIL_HOST_USER}>",

                # Email qui reçoit le message
                to=[settings.EMAIL_HOST_USER],

                # Réponse directe vers l'utilisateur
                reply_to=[email_utilisateur],
            )

            # Envoi du mail admin
            email_admin.send()

            # =========================
            # Réponse automatique à l'utilisateur
            # =========================
            message_confirmation = f"""
Bonjour {nom},

Votre message a été transmis avec succès à la Caisse Sociale DIRTECH.

Notre équipe l'étudiera dans les plus brefs délais et vous recevrez une réponse prochainement.

Merci pour votre confiance.

Cordialement,
Administration de la Caisse Sociale DIRTECH +221 775470087 / 785391575
"""

            confirmation = EmailMessage(
                subject="Confirmation de réception - Caisse Sociale DIRTECH",
                body=message_confirmation,

                from_email=f"Caisse Sociale DIRTECH <{settings.EMAIL_HOST_USER}>",

                # Mail du visiteur
                to=[email_utilisateur],
            )

            # Envoi de la confirmation
            confirmation.send()

            # =========================
            # Succès
            # =========================
            success = True

            # Réinitialisation du formulaire
            form = ContactForm()

    else:

        form = ContactForm()

    return render(request, 'nous_ecrire.html', {
        'form': form,
        'success': success
    })

def objectif(request):
    return render(request, 'app_social/objectif.html')

def reglement_interieur(request):
    return render(request, 'app_social/reglement_interieur.html')

def tableau_bord(request):
    return render(request, 'app_social/tableau_bord.html')

def le_bureau(request):
    return render(request, 'app_social/le_bureau.html')

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'

@login_required
def admin_dashboard(request):

    # =========================
    # 📊 STATISTIQUES PRÊTS
    # =========================
    total_prets = DemandePret.objects.count()

    montant_moyen_pret = DemandePret.objects.filter(
        statut='Validé'
    ).aggregate(moyenne=Avg('montant'))['moyenne'] or 0

    total_rembourses = Versement.objects.aggregate(
        total=Sum('montant')
    )['total'] or 0

    total_prets_valides = DemandePret.objects.filter(
        statut='Validé'
    ).aggregate(total=Sum('montant'))['total'] or 0

    taux_remboursement = (
        (total_rembourses / total_prets_valides) * 100
        if total_prets_valides > 0 else 0
    )

    # =========================
    # 📈 CROISSANCE
    # =========================
    today = date.today()

    somme_this_month = CotisationMensuelle.objects.filter(
        date__month=today.month,
        date__year=today.year
    ).aggregate(total=Sum('montant'))['total'] or 0

    last_month_date = today.replace(day=1) - timedelta(days=1)

    somme_last_month = CotisationMensuelle.objects.filter(
        date__month=last_month_date.month,
        date__year=last_month_date.year
    ).aggregate(total=Sum('montant'))['total'] or 0

    croissance = (
        ((somme_this_month - somme_last_month) / somme_last_month) * 100
        if somme_last_month > 0 else 0
    )

    # =========================
    # 💰 SOLDE
    # =========================
    caisse = ParametreCaisse.objects.first()
    solde_net = caisse.solde_net if caisse else 0

    # =========================
    # 🔔 NOTIFICATIONS
    # =========================
    prets_en_attente = DemandePret.objects.filter(statut="En attente").count()

    cotisations_en_retard = 0

    aujourd_hui = timezone.now().date()

    prochain_remb = DemandePret.objects.filter(
        statut="Validé",
        date_fin__gte=aujourd_hui
    ).order_by('date_fin').first()

    # =========================
    # 📊 GRAPH COTISATIONS
    # =========================
    cotisations = CotisationMensuelle.objects.all()

    mois_data = defaultdict(int)

    # 🔥 éviter bug langue serveur
    try:
        locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
    except:
        pass

    for c in cotisations:
        if c.date:
            mois = c.date.strftime("%B").capitalize()
            mois_data[mois] += 1

    ordre_mois = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    labels = []
    values = []

    for m in ordre_mois:
        labels.append(m)
        values.append(mois_data.get(m, 0))

    # =========================
    # 🧠 FORMULAIRE
    # =========================
    form_adherant = AdherantForm()

    # =========================
    # 🔥 DEPENDANCES
    # =========================
    SOUS_DEPARTEMENTS_MAP = {
        'default': ['LE CAP', 'CCRO', 'PASTEUR', 'DANTEC'],
        'Dep_DELTA': ['DELTA'],
    }

    # =========================
    # 📦 CONTEXT FINAL
    # =========================
    context = {
        # 📊 stats
        'total_prets': total_prets,
        'taux_remboursement': round(taux_remboursement, 2),
        'montant_moyen_pret': round(montant_moyen_pret, 0),
        'croissance': round(croissance, 2),
        'solde_net': round(solde_net, 0),

        # 🔔 notif
        'prets_en_attente': prets_en_attente,
        'cotisations_en_retard': cotisations_en_retard,
        'prochain_remb': prochain_remb,

        # 🧠 form
        'form_adherant': form_adherant,

        # 📊 GRAPH (IMPORTANT FIX)
        'chart_labels': json.dumps(labels),
        'chart_values': json.dumps(values),

        # 🔥 JS
        'sous_dep_map': json.dumps(SOUS_DEPARTEMENTS_MAP),
    }

    return render(request, 'app_social/admin_dashboard.html', context)

# ✅ Prêts — voir pour tous, valider pour Président uniquement
@role_required('President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte')
def liste_demandes_pret(request):
    demandes = DemandePret.objects.all().order_by('-date_demande')
    return render(request, 'app_social/liste_demandes_pret.html', {'demandes': demandes})

@role_required('President')
def valider_pret(request, demande_id):
    demande = get_object_or_404(DemandePret, id=demande_id)
    demande.statut = 'Validé'
    demande.save()
    messages.success(request, "✅ Demande validée.")
    return redirect('liste_demandes_pret')

@role_required('President')
def rejeter_pret(request, demande_id):
    demande = get_object_or_404(DemandePret, id=demande_id)
    demande.statut = 'Rejeté'
    demande.save()
    messages.warning(request, "❌ Demande rejetée.")
    return redirect('liste_demandes_pret')


def is_admin(user):
    return user.is_authenticated and user.is_staff

@role_required('President')
def changer_statut_pret(request, demande_id):
    demande = get_object_or_404(DemandePret, id=demande_id)

    if request.method == "POST":
        action = request.POST.get('action')

        if action == "valider":
            demande.statut = "Validé"
            demande.save()
            messages.success(request, "✅ Prêt validé")

        elif action == "refuser":
            demande.statut = "Refusé"
            demande.save()
            messages.warning(request, "❌ Prêt refusé")

        elif action == "attente":
            demande.statut = "En attente"
            demande.save()
            messages.info(request, "⏳ Remis en attente")

        # 🔥 SUPPRESSION
        elif action == "supprimer":
            demande.delete()
            messages.success(request, "🗑️ Prêt supprimé avec succès")

        return redirect('liste_demandes_pret')


def liste_prets(request):
    liste_prets = Pret.objects.all()
    prets_valides = Pret.objects.filter(statut='Validée').select_related('adherant')
    
    context = {
        'liste_prets': liste_prets,
        'prets_valides': prets_valides
    }
    return render(request, 'app_social/liste_prets.html', context)

    # Configurez le logging pour debug
logger = logging.getLogger(__name__)

@role_required('President', 'Tresorier', 'Adjoint_Tresorier')
def ajouter_cotisation(request):
    if request.method == 'POST':
        logger.info("=== 🔍 DEBUG COTISATION POST ===")
        logger.info(f"POST data: {request.POST}")

        adherant_id = request.POST.get('adherant_id')
        montant = request.POST.get('montant', 5000)
        mois = request.POST.get('mois')
        annee = request.POST.get('annee')
        date_cotisation = request.POST.get('date_cotisation')
        next_url = request.POST.get('next', 'cotisations_par_mois')  # ← page de retour

        logger.info(f"Adhérant ID: {adherant_id}")
        logger.info(f"Mois reçu: '{mois}' (type: {type(mois)})")
        logger.info(f"Année: {annee}")
        logger.info(f"Date: {date_cotisation}")
        logger.info("============================")

        if not adherant_id:
            messages.error(request, "❌ Aucun adhérent sélectionné.")
            return redirect(next_url)

        if not mois:
            messages.error(request, "❌ Le mois est obligatoire.")
            return redirect(next_url)

        adherant = get_object_or_404(Adherant, id=adherant_id)

        if adherant.statut != 'actif':
            messages.error(
                request,
                f"❌ Impossible d'ajouter une cotisation : {adherant.nom} {adherant.prenom} est un ancien adhérent."
            )
            return redirect(next_url)

        existe_deja = CotisationMensuelle.objects.filter(
            adherant=adherant,
            mois=mois,
            annee=annee
        ).exists()

        if existe_deja:
            messages.warning(request, f"⚠️ Cotisation déjà enregistrée pour {adherant.nom} {adherant.prenom} - {mois} {annee}")
            return redirect(next_url)

        try:
            CotisationMensuelle.objects.create(
                adherant=adherant,
                montant=5000,
                mois=mois,
                annee=int(annee),
                date_cotisation=date_cotisation,
            )

            adherant.date_dernier_cotisation = date.today()
            adherant.save()

            messages.success(request, f"✅ Cotisation ajoutée pour {adherant.nom} {adherant.prenom} - {mois} {annee}")
            logger.info(f"✅ Cotisation créée avec succès - Mois: {mois}")

        except Exception as e:
            logger.error(f"❌ Erreur lors de la création de la cotisation: {str(e)}")
            messages.error(request, f"❌ Erreur: {str(e)}")

        return redirect(next_url)

    return redirect('dashboard')



ROLES_BUREAU = ['President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte']

@login_required(login_url='login_admin')
def dashboard(request):
    est_bureau = request.user.groups.filter(name__in=ROLES_BUREAU).exists()
    
    if not request.user.is_staff and not est_bureau:
        return redirect('login_adherent')

    # ------------------------------------------------------------------------
    # Dates de référence
    # ------------------------------------------------------------------------
    maintenant = datetime.now()
    aujourd_hui = date.today()
    mois = maintenant.month
    annee = maintenant.year
    mois_precedent = maintenant - relativedelta(months=1)

    # ------------------------------------------------------------------------
    # Adhérents
    # ------------------------------------------------------------------------
    adherants = Adherant.objects.all().order_by('nom', 'prenom')
    total_adherents = adherants.count()

    adherents_cotisants_mois = CotisationMensuelle.objects.filter(
        date_cotisation__year=annee,
        date_cotisation__month=mois
    ).values('adherant').distinct().count()

    # ------------------------------------------------------------------------
    # Cotisations
    # ------------------------------------------------------------------------
    cot_mois_en_cours = CotisationMensuelle.objects.filter(
        date_cotisation__year=annee,
        date_cotisation__month=mois
    )

    total_montant = cot_mois_en_cours.aggregate(Sum('montant'))['montant__sum'] or 0
    total_montant_mois = total_montant

    fond_initial_obj = ParametreCaisse.objects.first()
    fond = fond_initial_obj.fond_initial if fond_initial_obj else 0

    montant_total_cotisations = CotisationMensuelle.objects.aggregate(Sum('montant'))['montant__sum'] or 0

    # IDs des cotisants du mois en cours
    cotisants_ids = CotisationMensuelle.objects.filter(
        date_cotisation__year=annee,
        date_cotisation__month=mois
    ).values_list('adherant_id', flat=True).distinct()

    # Adhérents en retard
    adherents_en_retard = Adherant.objects.exclude(id__in=cotisants_ids)
    cotisations_en_retard = adherents_en_retard.count()

    # Croissance mensuelle
    cotisations_mois_actuel = Cotisation.objects.filter(
        date_cotisation__year=annee,
        date_cotisation__month=mois
    )
    total_cotisation_mois = cotisations_mois_actuel.aggregate(Sum('montant'))['montant__sum'] or 0

    cotisations_mois_precedent = Cotisation.objects.filter(
        date_cotisation__year=mois_precedent.year,
        date_cotisation__month=mois_precedent.month
    )
    total_cotisation_precedent = cotisations_mois_precedent.aggregate(Sum('montant'))['montant__sum'] or 0

    croissance = 0
    if total_cotisation_precedent > 0:
        croissance = ((total_cotisation_mois - total_cotisation_precedent) / total_cotisation_precedent) * 100

    cot_par_mois = CotisationMensuelle.objects.annotate(
        mois_trunc=TruncMonth('date_cotisation')
    ).values('mois_trunc').annotate(total=Sum('montant')).order_by('mois_trunc')
    labels = [e['mois_trunc'].strftime('%b %Y') for e in cot_par_mois]
    data = [float(e['total']) for e in cot_par_mois]

    # ------------------------------------------------------------------------
    # Prêts & Remboursements
    # ------------------------------------------------------------------------
    total_prets_valides = Pret.objects.filter(statut='Validé').aggregate(Sum('montant'))['montant__sum'] or 0
    montant_moyen_pret = Pret.objects.filter(statut='Validé').aggregate(Avg('montant'))['montant__avg'] or 0
    total_remboursements = Versement.objects.aggregate(total=Sum('montant'))['total'] or Decimal(0)
    taux_remboursement = (total_remboursements / total_prets_valides) * 100 if total_prets_valides > 0 else 0

    remboursements = PaiementPret.objects.values('pret').annotate(total_rembourse=Sum('montant'))
    remboursement_dict = {r['pret']: r['total_rembourse'] for r in remboursements}
    prets_valides = Pret.objects.filter(statut='Validé')

    total_restant_a_rembourser = sum(
        float(pret.montant) - float(remboursement_dict.get(pret.id, 0)) for pret in prets_valides
    )

    # ------------------------------------------------------------------------
    # Notifications & Prochains paiements
    # ------------------------------------------------------------------------
    prets_en_attente = DemandePret.objects.filter(statut='En attente').count()

    cotisants_ids2 = Cotisation.objects.annotate(
        mois_cotisation=ExtractMonth('date_cotisation'),
        annee_cotisation=ExtractYear('date_cotisation')
    ).filter(mois_cotisation=mois, annee_cotisation=annee).values_list('adherant_id', flat=True).distinct()

    adherents_a_jour = Adherant.objects.filter(id__in=cotisants_ids2)
    adherents_en_retard = Adherant.objects.exclude(id__in=cotisants_ids2)

    dans_7_jours = maintenant + timedelta(days=7)
    prochain_remb = PaiementPret.objects.filter(
        date_paiement__range=(maintenant, dans_7_jours)
    ).order_by('date_paiement')[:5]

    # ------------------------------------------------------------------------
    # Solde net
    # ------------------------------------------------------------------------
    fond_initial = ParametreCaisse.objects.aggregate(Sum("fond_initial"))["fond_initial__sum"] or 0
    montant_total_cotisations = CotisationMensuelle.objects.aggregate(Sum("montant"))["montant__sum"] or 0
    total_remboursements = Versement.objects.aggregate(Sum("montant"))["montant__sum"] or 0
    total_prets_valides = DemandePret.objects.filter(statut="Validé").aggregate(Sum("montant"))["montant__sum"] or 0
    total_decaissements = Decaissement.objects.aggregate(Sum("montant"))["montant__sum"] or 0

    solde_net = (
        fond_initial
        + montant_total_cotisations
        + total_remboursements
        - total_prets_valides
        - total_decaissements
    )

    # ------------------------------------------------------------------------
    # Activités récentes — 5 dernières cotisations
    # ------------------------------------------------------------------------
    dernieres_cotisations = CotisationMensuelle.objects.select_related('adherant').order_by('-date_cotisation')[:5]

    activites_recentes = []
    for cot in dernieres_cotisations:
        activites_recentes.append({
            'type': 'cotisation',
            'icone': 'bi-cash-stack',
            'couleur': 'success',
            'texte': f"{cot.adherant.nom} {cot.adherant.prenom}",
            'detail': f"{cot.mois} {cot.annee} — {int(cot.montant):,} FCFA".replace(',', ' '),
            'date': cot.date_cotisation,
        })

    # ------------------------------------------------------------------------
    # 3 derniers décaissements récents
    # ------------------------------------------------------------------------
    derniers_decaissements = Decaissement.objects.select_related('adherant').order_by('-date')[:3]

    decaissements_recents = []
    for dec in derniers_decaissements:
        decaissements_recents.append({
            'nom': f"{dec.adherant.nom} {dec.adherant.prenom}",
            'motif': dec.motif,
            'montant': int(dec.montant),
            'date': dec.date,
        })

    # ------------------------------------------------------------------------
    # Logs et contexte
    # ------------------------------------------------------------------------
    recent_logs = LogEntry.objects.select_related('user').order_by('-action_time')[:5]
    mois_list = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                 "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

    context = {
        'total_adherents': total_adherents,
        'total_cotisations': total_cotisation_mois,
        'total_prets': total_prets_valides,
        'cotisations_mois': cot_mois_en_cours.count(),
        'total_montant': total_montant,
        'labels': labels,
        'data': data,
        'recent_logs': recent_logs,
        'mois_list': mois_list,
        'adherants': adherants,
        'now': maintenant,
        'croissance': round(croissance, 2),
        'montant_moyen_pret': round(montant_moyen_pret, 0),
        'taux_remboursement': round(taux_remboursement, 2),
        'solde_net': solde_net,
        'total_restant_a_rembourser': total_restant_a_rembourser,
        'nb_cotisants_mois': adherents_cotisants_mois,
        'total_montant_mois': total_montant_mois,
        'prets_en_attente': prets_en_attente,
        'cotisations_en_retard': cotisations_en_retard,
        'prochain_remb': prochain_remb,
        'adherents_a_jour': adherents_a_jour,
        'adherents_en_retard': adherents_en_retard,
        'total_decaissements': total_decaissements,
        'activites_recentes': activites_recentes,
        'decaissements_recents': decaissements_recents,
        'est_bureau': est_bureau,
        'role_actuel': request.user.groups.first().name if request.user.groups.exists() else None,
    }

    return render(request, "admin_dashboard.html", context)


# ✅ APRÈS — tous les rôles bureau y ont accès
@role_required('President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte')
def solde_net(request):
    # Récupération du fond initial
    parametres = ParametreCaisse.objects.first()
    fond_initial = parametres.fond_initial if parametres else 0

    # Calcul des totaux
    montant_total_cotisations = CotisationMensuelle.objects.aggregate(
        total=Sum('montant'))['total'] or 0

    total_remboursements = PaiementPret.objects.aggregate(
        total=Sum('montant'))['total'] or 0

    total_prets_valides = Pret.objects.filter(
        statut='Validé').aggregate(total=Sum('montant'))['total'] or 0

    # Solde net
    solde_net = fond_initial + montant_total_cotisations + total_remboursements - total_prets_valides

    return render(request, 'notifications/solde_net.html', {'solde_net': solde_net})

@role_required('President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte')
def prets_en_attente(request):
    prets_en_attente = DemandePret.objects.filter(statut='en_attente').order_by('-date_demande')

    return render(request, 'notifications/prets_en_attente.html', {'prets_en_attente': prets_en_attente})

@role_required('President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte')
def cotisations_en_retard(request):
    aujourd_hui = date.today()
    mois = aujourd_hui.month
    annee = aujourd_hui.year

    # Tous les adhérents
    adherents = Adherant.objects.all()

    # Ceux qui ont déjà cotisé ce mois
    cotisants_ids = CotisationMensuelle.objects.filter(
        date_cotisation__year=annee,
        date_cotisation__month=mois
    ).values_list('adherant_id', flat=True).distinct()

    # Ceux qui ne sont PAS dans la liste des cotisants
    non_cotisants = adherents.exclude(id__in=cotisants_ids).order_by('nom', 'prenom')

    context = {
        "non_cotisants": non_cotisants
    }
    return render(request, "notifications/cotisations_en_retard.html", context)


def prochain_remboursement(request):
    prochain_remb = PaiementPret.objects.filter(
        date_paiement__gte=date.today()
    ).order_by('date_paiement').first()    
    context = {
        'prochain_remb': prochain_remb,
    }
    return render(request, 'notifications/prochain_remboursement.html', context)


def adherents_a_jour(request):
    mois_actuel = now().month
    annee_actuelle = now().year

    cotisants_ids = CotisationMensuelle.objects.filter(
        date_cotisation__month=mois_actuel,
        date_cotisation__year=annee_actuelle
    ).values_list('adherant_id', flat=True)

    adherents_a_jour = Adherant.objects.filter(id__in=cotisants_ids)

    context = {
        'adherents_a_jour': adherents_a_jour,
    }
    return render(request, 'notifications/adherents_a_jour.html', context)

def adherents_non_a_jour(request):
    mois_actuel = now().month
    annee_actuelle = now().year

    cotisants_ids = CotisationMensuelle.objects.filter(
        date_cotisation__month=mois_actuel,
        date_cotisation__year=annee_actuelle
    ).values_list('adherant_id', flat=True)

    adherents_non_a_jour = Adherant.objects.exclude(id__in=cotisants_ids)

    context = {
        'adherents_non_a_jour': adherents_non_a_jour,
    }
    return render(request, 'notifications/adherents_non_a_jour.html', context)


def adherents_en_retard(request):
    now = datetime.now()
    mois_actuel = now.month
    annee_actuelle = now.year

    adherents = Adherant.objects.all()
    adherents_en_retard = []

    for adherent in adherents:
        cotisation = CotisationMensuelle.objects.filter(
            adherant=adherent,
            date_cotisation__month=mois_actuel,
            date_cotisation__year=annee_actuelle
        ).first()
        if not cotisation:
            adherents_en_retard.append(adherent)

    return render(request, 'notifications/adherents_en_retard.html', {'adherents_en_retard': adherents_en_retard})

# ✅ Cotisations — tous sauf Vice-Président pour ajouter/supprimer
@role_required('President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte')
def cotisations_par_mois(request):
    """Vue corrigée pour la cohérence des noms de mois"""
    mois_choisi = request.GET.get('mois') or 'Janvier'
    annee_choisie = int(request.GET.get('annee') or datetime.now().year)
    search_query = request.GET.get('search', '').strip()

    # Cotisants existants pour le mois/année
    cotisants_ids = CotisationMensuelle.objects.filter(
        mois=mois_choisi,  # Utilise le nom du mois
        annee=annee_choisie
    ).values_list('adherant_id', flat=True)

    # Exclure ces adhérents
    adherants = Adherant.objects.exclude(id__in=cotisants_ids).order_by('nom', 'prenom')

    cotisations = CotisationMensuelle.objects.filter(
        mois=mois_choisi,  # Utilise le nom du mois
        annee=annee_choisie
    ).select_related('adherant').order_by('adherant__nom', 'adherant__prenom')

    if search_query:
        cotisations = cotisations.filter(
            Q(adherant__nom__icontains=search_query) |
            Q(adherant__prenom__icontains=search_query)
        )

    total_montant = cotisations.aggregate(total=Sum('montant'))['total'] or 0

    # Options de mois cohérentes
    mois_options = [
        ('Janvier', 'Janvier'), ('Février', 'Février'), ('Mars', 'Mars'), ('Avril', 'Avril'),
        ('Mai', 'Mai'), ('Juin', 'Juin'), ('Juillet', 'Juillet'), ('Aout', 'Aout'),
        ('Septembre', 'Septembre'), ('Octobre', 'Octobre'), ('Novembre', 'Novembre'), ('Decembre', 'Decembre')
    ]

    context = {
        'cotisations': cotisations,
        'mois_options': mois_options,
        'mois_choisi': mois_choisi,
        'annee_choisie': annee_choisie,
        'search_query': search_query,
        'total_montant': total_montant,
        'adherants': adherants, 
    }

    return render(request, 'admin/cotisations_par_mois.html', context)


def telecharger_cotisations_excel(request):
    mois = request.GET.get('mois')
    annee = request.GET.get('annee')

    cotisations = CotisationMensuelle.objects.filter(mois=mois, annee=annee)

    wb = Workbook()
    ws = wb.active
    ws.title = f'Cotisations {mois} {annee}'

    headers = ['Nom', 'Prénom', 'Montant (FCFA)', 'Date de cotisation']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for cot in cotisations:
        ws.append([
            cot.adherant.nom,
            cot.adherant.prenom,
            cot.montant,
            cot.date_cotisation.strftime("%d/%m/%Y")
        ])

    column_widths = [20, 20, 18, 25]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'cotisations_{mois}_{annee}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response


from django.shortcuts import get_object_or_404, render
from datetime import date
from .models import Adherant, Pret, CotisationMensuelle

def fiche_adherant(request, adherant_id):
    adherant = get_object_or_404(Adherant, id=adherant_id)

    today = date.today()

    # ===================== PRÊT EN COURS =====================
    a_pret_en_cours = Pret.objects.filter(
        adherant=adherant,
        statut__in=['En cours', 'Validé']  # On prend aussi les prêts validés
    ).exists()

    # ===================== COTISATION DU MOIS COURANT =====================
    # Mapping du mois en français
    mois_mapping = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
        5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
        9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
    }

    mois_actuel_str = mois_mapping.get(today.month, "")

    a_jour_cotisation = CotisationMensuelle.objects.filter(
        adherant=adherant,
        mois=mois_actuel_str,      # Important : mois en texte
        annee=today.year
    ).exists()

    # ===================== HISTORIQUE DES COTISATIONS =====================
    dernieres_cotisations = CotisationMensuelle.objects.filter(
        adherant=adherant
    ).select_related('adherant').order_by('-annee', '-date_cotisation')[:12]  # 12 derniers mois

    context = {
        'adherant': adherant,
        'a_pret_en_cours': a_pret_en_cours,
        'a_jour_cotisation': a_jour_cotisation,
        'dernieres_cotisations': dernieres_cotisations,
        'mois_actuel': mois_actuel_str,
        'annee_actuelle': today.year,
    }

    return render(request, 'app_social/fiche_adherant.html', context)


def fiche_adherant_pdf(request, adherant_id):
    adherant = get_object_or_404(Adherant, id=adherant_id)
    today = date.today()

    # Données
    a_pret_en_cours = Pret.objects.filter(
        adherant=adherant, statut__in=['En cours', 'Validé']
    ).exists()

    mois_mapping = {1:"Janvier",2:"Février",3:"Mars",4:"Avril",5:"Mai",6:"Juin",
                    7:"Juillet",8:"Août",9:"Septembre",10:"Octobre",11:"Novembre",12:"Décembre"}
    mois_str = mois_mapping.get(today.month, "")

    a_jour_cotisation = CotisationMensuelle.objects.filter(
        adherant=adherant, mois=mois_str, annee=today.year
    ).exists()

    dernieres_cotisations = CotisationMensuelle.objects.filter(
        adherant=adherant
    ).order_by('-annee', '-date_cotisation')[:20]

    # ===================== PDF =====================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Fiche_Adherent_{adherant.nom}_{adherant.prenom}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, 
                            rightMargin=2*cm, leftMargin=2*cm, 
                            topMargin=2*cm, bottomMargin=2*cm)

    elements = []
    styles = getSampleStyleSheet()

    # Style personnalisé
    center_style = ParagraphStyle('center', alignment=1, fontSize=11, spaceAfter=6)

    # ==================== EN-TÊTE OFFICIEL ====================
    elements.append(Paragraph("<b>RÉPUBLIQUE DU SÉNÉGAL</b>", center_style))
    elements.append(Paragraph("<b>PRESIDENCE DE LA RÉPUBLIQUE</b>", center_style))
    elements.append(Paragraph("Délégation Générale au Renseignement National", center_style))
    elements.append(Paragraph("<b>DIRECTION TECHNIQUE</b>", center_style))
    elements.append(Paragraph("<b>CAISSE SOCIALE</b>", ParagraphStyle('title', alignment=1, fontSize=14, spaceAfter=20)))

    elements.append(Spacer(1, 15))

    # Numérotation Ref/CSDRN/FA/XXXX
    ref_number = f"Ref/CSDRN/FA/{adherant.id:04d}"   # Tu peux améliorer la numérotation plus tard
    elements.append(Paragraph(f"<b>{ref_number}</b>", ParagraphStyle('ref', alignment=2, fontSize=10)))

    elements.append(Spacer(1, 25))

    # Titre du document
    elements.append(Paragraph("FICHE INDIVIDUELLE D'ADHÉRENT", 
                             ParagraphStyle('main_title', alignment=1, fontSize=16, spaceAfter=20)))

    # ==================== INFORMATIONS ====================
    info_data = [
        ["Nom et Prénom", f"{adherant.nom} {adherant.prenom}"],
        ["Téléphone", adherant.telephone],
        ["Email", adherant.email or "Non renseigné"],
        ["Adresse", adherant.adresse_rue],
        ["Département", adherant.departement],
        ["Sous-département", adherant.sous_departement or "-"],
        ["Statut", adherant.statut.upper()],
    ]

    info_table = Table(info_data, colWidths=[6*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))

    # Situation Financière
    elements.append(Paragraph("SITUATION FINANCIÈRE", styles['Heading3']))
    situation_data = [
        ["Désignation", "Statut"],
        ["Prêt en cours", "OUI" if a_pret_en_cours else "NON"],
        [f"Cotisation {mois_str} {today.year}", "À JOUR" if a_jour_cotisation else "EN RETARD"],
    ]

    sit_table = Table(situation_data, colWidths=[8*cm, 8*cm])
    sit_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(sit_table)
    elements.append(Spacer(1, 25))

    # Historique des Cotisations
    elements.append(Paragraph("HISTORIQUE DES COTISATIONS", styles['Heading3']))

    table_data = [["Date", "Mois", "Année", "Montant (FCFA)"]]
    for cot in dernieres_cotisations:
        table_data.append([
            cot.date_cotisation.strftime("%d/%m/%Y"),
            cot.mois,
            str(cot.annee),
            f"{int(cot.montant):,} FCFA".replace(',', ' ')
        ])

    if not dernieres_cotisations:
        table_data.append(["", "Aucune cotisation enregistrée", "", ""])

    cot_table = Table(table_data, colWidths=[4*cm, 4*cm, 3*cm, 5*cm])
    cot_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    elements.append(cot_table)

    # Pied de page
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(f"Document généré le {today.strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))

    doc.build(elements)
    return response

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Adherant

def upload_photo_adherant(request, adherant_id):
    adherant = get_object_or_404(Adherant, id=adherant_id)

    if request.method == 'POST' and request.FILES.get('photo'):
        photo = request.FILES['photo']

        if photo.content_type not in ['image/jpeg', 'image/png', 'image/jpg']:
            messages.error(request, "Seuls les fichiers JPG et PNG sont autorisés.")
            return redirect('fiche_adherant', adherant_id=adherant.id)

        adherant.photo = photo
        adherant.save()

        messages.success(request, "✅ Photo mise à jour avec succès !")
    else:
        messages.error(request, "Aucune photo n'a été envoyée.")

    return redirect('fiche_adherant', adherant_id=adherant.id)

@role_required('President', 'Tresorier', 'Adjoint_Tresorier')
def supprimer_cotisation(request, cotisation_id):
    cotisation = get_object_or_404(CotisationMensuelle, id=cotisation_id)
    cotisation.delete()
    messages.success(request, "La cotisation a été supprimée avec succès.")
    return redirect('cotisations_par_mois')  # ou la vue où tu veux revenir

def upload_photo_adherant(request, adherant_id):
    adherant = get_object_or_404(Adherant, id=adherant_id)
    if request.method == 'POST' and request.FILES.get('photo'):
        adherant.photo = request.FILES['photo']
        adherant.save()
    return redirect('fiche_adherant', adherant_id=adherant.id)

# ✅ Paiement prêts — Président, Trésorier, Adjoint
@role_required('President', 'Tresorier', 'Adjoint_Tresorier')
def paiement_pret(request):
    """
    Vue pour afficher les prêts validés et permettre l'ajout de versements
    """
    # Récupérer tous les prêts avec le statut 'Validé' (attention à la casse)
    prets_valides = DemandePret.objects.filter(
        statut__iexact='validé'  # iexact pour ignorer la casse
    ).select_related('adherant').prefetch_related('versements')
    
    # Debug: afficher les statuts disponibles
    if not prets_valides.exists():
        # Vérifier quels statuts existent réellement
        statuts_existants = DemandePret.objects.values_list('statut', flat=True).distinct()
        messages.info(request, f"Statuts disponibles: {list(statuts_existants)}")
    
    context = {
        'prets_valides': prets_valides,
    }
    
    return render(request, 'app_social/paiement_pret.html', context)


@role_required('President', 'Tresorier', 'Adjoint_Tresorier')
def enregistrer_versement(request):
    if request.method == 'POST':
        pret_id = request.POST.get('pret_id')
        montant_verse = request.POST.get('montant_verse')
        next_url = request.POST.get('next', 'paiement_pret')  # ← page de retour

        if not pret_id or not montant_verse:
            messages.error(request, "Données manquantes pour l'enregistrement du versement.")
            return redirect(next_url)

        try:
            pret = get_object_or_404(DemandePret, id=pret_id)
            montant_decimal = Decimal(str(montant_verse))

            if montant_decimal <= 0:
                messages.error(request, "Le montant doit être supérieur à 0.")
                return redirect(next_url)

            if montant_decimal > pret.reste_a_payer:
                messages.error(request, f"Le montant ({montant_decimal} FCFA) dépasse le reste à payer ({pret.reste_a_payer} FCFA).")
                return redirect(next_url)

            Versement.objects.create(pret=pret, montant=montant_decimal)
            total_verse = Versement.objects.filter(pret=pret).aggregate(total=Sum('montant'))['total'] or 0

            if total_verse >= pret.montant:
                pret.statut = 'Remboursé'
                pret.save()
                messages.success(request, "✅ Versement enregistré. Le prêt est maintenant entièrement remboursé !")
            else:
                messages.success(
                    request,
                    f"✅ Versement de {montant_decimal} FCFA enregistré avec succès pour {pret.adherant.nom} {pret.adherant.prenom}. "
                    f"Reste à payer: {pret.reste_a_payer} FCFA"
                )

        except ValueError as e:
            messages.error(request, f"Montant invalide: {e}")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'enregistrement: {str(e)}")
            import logging
            logging.error(f"Erreur versement - Prêt ID: {pret_id}, Montant: {montant_verse}, Erreur: {str(e)}")

    return redirect(next_url)



# Dictionnaire des montants fixes
MONTANT_CHOICES = {
    "Bapteme": "50000",
    "Mariage": "75000",
    "Maladie": "30000",
    "Décès": "150000",
    "Depart Retraite/Affectation": "100000",
}


MONTANT_CHOICES = {
    "Bapteme": "50000",
    "Mariage": "75000",
    "Maladie": "30000",
    "Décès": "150000",
    "Depart Retraite/Affectation": "100000",
}

from django.utils import timezone

# ✅ Décaissements — Président et Trésorier
@role_required('President', 'Tresorier')
def handle_decaissement(request, motif):
    if request.method == "POST":
        form = DecaissementForm(request.POST, request.FILES)
        if form.is_valid():
            decaissement = form.save(commit=False)

            # Motif
            decaissement.motif = motif  

            # Montant automatique
            if motif in MONTANT_CHOICES:
                decaissement.montant = Decimal(MONTANT_CHOICES[motif])

            if motif == "Panier Ndogou/Noël":
                decaissement.montant = form.cleaned_data.get("montant") or 0

            decaissement.save()

            # 🔹 Archiver et désactiver l'adhérent si départ retraite/affectation
            if motif == "Depart Retraite/Affectation" and decaissement.adherant:
                adherent = decaissement.adherant

                # Archiver
                AncienAdherent.objects.create(
                    adherant=adherent,
                    date_depart=timezone.now(),
                    motif_depart=motif
                )

                # Désactiver l'adhérent (au lieu de supprimer)
                adherent.actif = False
                adherent.save()

                messages.info(
                    request,
                    f"L’adhérent {adherent.nom} {adherent.prenom} a été archivé et désactivé."
                )

            # Message succès avec lien PDF
            messages.success(
                request,
                mark_safe(
                    f"✅ Décaissement '{motif}' enregistré avec succès ! "
                    f"<a href='{reverse('download_decaissement_pdf', args=[decaissement.id])}' "
                    f"class='btn btn-sm btn-primary ms-2' target='_blank'>"
                    f"Télécharger la fiche</a>"
                )
            )

            return redirect(request.path)

    else:
        form = DecaissementForm()

    return render(request, "decaissements/decaissement_form.html", {
        "form": form,
        "motif": motif,
    })


# Vues spécifiques par motif
def decaissement_bapteme(request):
    return handle_decaissement(request, "Bapteme")

def decaissement_mariage(request):
    return handle_decaissement(request, "Mariage")

def decaissement_maladie(request):
    return handle_decaissement(request, "Maladie")

def decaissement_deces(request):
    return handle_decaissement(request, "Décès")

def decaissement_panier(request):
    return handle_decaissement(request, "Panier Ndogou/Noël")

def decaissement_depart(request):
    return handle_decaissement(request, "Depart Retraite/Affectation")

# Vue pour afficher la fiche en HTML
def fiche_decaissement_view(request, decaissement_id):
    decaissement = get_object_or_404(Decaissement, id=decaissement_id)
    return render(request, "decaissements/fiche_decaissement.html", {
        "decaissement": decaissement
    })

def fiche_decaissement(request, decaissement_id):
    decaissement = get_object_or_404(Decaissement, id=decaissement_id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="fiche_decaissement_{decaissement.id}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # --- ENTÊTE ---
    p.setFont("Helvetica-Bold", 12)
    p.drawCentredString(width / 2, height - 50, "REPUBLIQUE DU SENEGAL")
    p.setFont("Helvetica-Bold", 11)
    p.drawCentredString(width / 2, height - 70, "PRESIDENCE DE LA REPUBLIQUE")
    p.setFont("Helvetica-Bold", 10)
    p.drawCentredString(width / 2, height - 90, "Délégation Générale au Renseignement National")
    p.setFont("Helvetica-Bold", 11)
    p.drawCentredString(width / 2, height - 110, "DIRECTION TECHNIQUE")
    p.setFont("Helvetica-Oblique", 11)
    p.drawCentredString(width / 2, height - 130, "CAISSE SOCIALE")

    # --- TITRE ---
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width / 2, height - 180, "FICHE DE DÉCAISSEMENT")

    # --- TABLEAU DES INFOS ---
    data = [
        ["Motif", decaissement.motif],
        ["Montant", f"{decaissement.montant} FCFA"],
        ["Date", str(decaissement.date)],
        ["Adhérent", f"{decaissement.adherant.nom} {decaissement.adherant.prenom}"],
        ["Téléphone", decaissement.adherant.telephone],
    ]

    table = Table(data, colWidths=[150, 300])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),  # Colonne 1 en gris clair
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),  # Première ligne en gras
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
    ]))

    table.wrapOn(p, width, height)
    table.drawOn(p, 50, height - 350)

    # --- SIGNATURES ---
    data_signatures = [
        ["Le Trésorier", "Le Président"],
        ["\n\n\n_____________________", "\n\n\n_____________________"]
    ]

    table_sign = Table(data_signatures, colWidths=[width/2 - 60, width/2 - 60])
    table_sign.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 20),
    ]))

    table_sign.wrapOn(p, width, height)
    table_sign.drawOn(p, 50, height - 500)

    p.showPage()
    p.save()
    return response

# ✅ Rapports/Décaissements — tous
@role_required('President', 'Vice_President', 'Tresorier', 'Adjoint_Tresorier', 'Commissaire_Compte')
def liste_decaissements(request):
    decaissements = Decaissement.objects.all()
    return render(request, "liste_decaissements.html", {"decaissements": decaissements})


def download_decaissement_pdf(request, decaissement_id):
    decaissement = get_object_or_404(Decaissement, id=decaissement_id)
    adherant = decaissement.adherant
    date_edition = datetime.now().strftime('%d/%m/%Y à %H:%M')

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # ============================================================
    # COULEURS
    # ============================================================
    bleu_fonce  = colors.HexColor('#0f172a')
    bleu_moyen  = colors.HexColor('#1e3a5f')
    bleu_clair  = colors.HexColor('#e8f0fe')
    or_couleur  = colors.HexColor('#c9a84c')
    gris_clair  = colors.HexColor('#f8fafc')
    gris_ligne  = colors.HexColor('#e2e8f0')
    blanc       = colors.white

    # ============================================================
    # FILIGRANE — discret, en arrière-plan
    # ============================================================
    p.saveState()
    p.setFillColor(colors.HexColor('#e2e8f0'))
    p.setFont("Helvetica-Bold", 40)
    p.translate(width / 2, 180)
    p.rotate(25)
    p.setFillAlpha(0.08)
    p.drawCentredString(0, 0, "CAISSE SOCIALE DIRTECH")
    p.restoreState()

    # ============================================================
    # BANDEAU HEADER
    # ============================================================
    p.setFillColor(bleu_fonce)
    p.rect(0, height - 110, width, 110, fill=1, stroke=0)

    p.setFillColor(or_couleur)
    p.rect(0, height - 113, width, 3, fill=1, stroke=0)

    p.setFillColor(blanc)
    p.setFont("Helvetica-Bold", 10)
    p.drawCentredString(width / 2, height - 28, "REPUBLIQUE DU SENEGAL")
    p.setFont("Helvetica", 9)
    p.drawCentredString(width / 2, height - 44, "PRESIDENCE DE LA REPUBLIQUE")
    p.drawCentredString(width / 2, height - 58, "Délégation Générale au Renseignement National")
    p.drawCentredString(width / 2, height - 72, "DIRECTION TECHNIQUE")
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(or_couleur)
    p.drawCentredString(width / 2, height - 90, "✦  CAISSE SOCIALE  ✦")

    # ============================================================
    # BLOC TITRE FICHE
    # ============================================================
    p.setFillColor(bleu_moyen)
    p.roundRect(40, height - 165, width - 80, 40, 6, fill=1, stroke=0)

    p.setFillColor(blanc)
    p.setFont("Helvetica-Bold", 13)
    p.drawCentredString(width / 2, height - 148,
        f"FICHE DE DÉCAISSEMENT  —  {decaissement.motif.upper()}")

    p.setFillColor(bleu_fonce)
    p.setFont("Helvetica", 8)
    p.drawRightString(width - 40, height - 178,
        f"N° REF : DEC-{decaissement.id:04d}  |  Date d'édition : {date_edition}")

    # ============================================================
    # SECTION — INFORMATIONS BÉNÉFICIAIRE
    # ============================================================
    y = height - 210

    p.setFillColor(bleu_clair)
    p.rect(40, y - 4, width - 80, 22, fill=1, stroke=0)
    p.setFillColor(bleu_moyen)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y + 4, "▌  INFORMATIONS DU BÉNÉFICIAIRE")

    y -= 30

    infos_benef = [
        ["Nom complet",   f"{adherant.nom.upper()} {adherant.prenom.capitalize()}"],
        ["Téléphone",     adherant.telephone or "—"],
        ["Département",   adherant.departement or "—"],
        ["Centre",        adherant.sous_departement or "—"],
        ["Adresse",       adherant.adresse_rue or "—"],
        ["Email",         adherant.email or "—"],
        ["Statut",        adherant.statut.upper() if adherant.statut else "—"],
    ]

    table_benef = Table(infos_benef, colWidths=[150, 355])
    table_benef.setStyle(TableStyle([
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [blanc, gris_clair]),
        ('GRID',        (0, 0), (-1, -1), 0.5, gris_ligne),
        ('FONTNAME',    (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME',    (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9.5),
        ('TEXTCOLOR',   (0, 0), (0, -1), bleu_moyen),
        ('TEXTCOLOR',   (1, 0), (1, -1), bleu_fonce),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
    ]))

    tw, th = table_benef.wrap(0, 0)
    table_benef.drawOn(p, 40, y - th)
    y = y - th - 20

    # ============================================================
    # SECTION — DÉTAILS DU DÉCAISSEMENT
    # ============================================================
    p.setFillColor(bleu_clair)
    p.rect(40, y - 4, width - 80, 22, fill=1, stroke=0)
    p.setFillColor(bleu_moyen)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y + 4, "▌  DÉTAILS DU DÉCAISSEMENT")

    y -= 30

    infos_dec = [
        ["Motif",        decaissement.motif],
        ["Date",         decaissement.date.strftime("%d/%m/%Y")],
        ["Description",  decaissement.description or "—"],
    ]

    table_dec = Table(infos_dec, colWidths=[150, 355])
    table_dec.setStyle(TableStyle([
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [blanc, gris_clair]),
        ('GRID',        (0, 0), (-1, -1), 0.5, gris_ligne),
        ('FONTNAME',    (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME',    (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9.5),
        ('TEXTCOLOR',   (0, 0), (0, -1), bleu_moyen),
        ('TEXTCOLOR',   (1, 0), (1, -1), bleu_fonce),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
    ]))

    tw, th = table_dec.wrap(0, 0)
    table_dec.drawOn(p, 40, y - th)
    y = y - th - 20

    # ============================================================
    # BLOC MONTANT ENCADRÉ
    # ============================================================
    p.setFillColor(bleu_moyen)
    p.roundRect(40, y - 55, width - 80, 50, 8, fill=1, stroke=0)

    p.setFillColor(or_couleur)
    p.setFont("Helvetica-Bold", 11)
    p.drawString(60, y - 22, "MONTANT ALLOUÉ :")

    montant_formate = f"{int(decaissement.montant):,}".replace(",", " ")
    p.setFillColor(blanc)
    p.setFont("Helvetica-Bold", 18)
    p.drawRightString(width - 60, y - 32, f"{montant_formate} FCFA")

    # ============================================================
    # SIGNATURES
    # ============================================================
    y_sign = 105

    p.setStrokeColor(gris_ligne)
    p.setLineWidth(0.5)
    p.line(40, y_sign + 75, width - 40, y_sign + 75)

    p.setFillColor(bleu_fonce)
    p.setFont("Helvetica-Bold", 9)
    p.drawString(60, y_sign + 60, "Le Trésorier")
    p.drawString(width - 200, y_sign + 60, "Le Président")

    p.setFont("Helvetica", 9)
    p.setFillColor(colors.HexColor('#94a3b8'))
    p.drawString(60, y_sign + 45, "Nom & Signature :")
    p.drawString(width - 200, y_sign + 45, "Nom & Signature :")

    p.setStrokeColor(gris_ligne)
    p.setLineWidth(0.5)
    p.rect(60, y_sign, 160, 40, fill=0, stroke=1)
    p.rect(width - 200, y_sign, 160, 40, fill=0, stroke=1)

    # ============================================================
    # PIED DE PAGE
    # ============================================================
    p.setFillColor(bleu_fonce)
    p.rect(0, 0, width, 38, fill=1, stroke=0)

    p.setFillColor(or_couleur)
    p.rect(0, 38, width, 2, fill=1, stroke=0)

    p.setFillColor(blanc)
    p.setFont("Helvetica", 7.5)
    p.drawCentredString(width / 2, 14,
        f"Document généré le {date_edition}  —  "
        f"Caisse Sociale DIRTECH  —  Réf. DEC-{decaissement.id:04d}")

    p.showPage()
    p.save()
    buffer.seek(0)

    filename = f"decaissement_{decaissement.motif.replace('/', '-')}_{decaissement.id:04d}.pdf"
    return FileResponse(buffer, as_attachment=False, filename=filename)

def liste_anciens_adherents(request):
    search_query = request.GET.get("search", "")
    departement_filter = request.GET.get("departement", "")
    sous_departement_filter = request.GET.get("sous_departement", "")

    # 🔹 On sélectionne uniquement les anciens adhérents liés à un adhérent existant
    anciens = AncienAdherent.objects.select_related('adherant').filter(adherant__isnull=False)

    # 🔹 Recherche par nom ou prénom
    if search_query:
        anciens = anciens.filter(
            Q(adherant__nom__icontains=search_query) |
            Q(adherant__prenom__icontains=search_query)
        )

    # 🔹 Filtre par département
    if departement_filter:
        anciens = anciens.filter(adherant__departement=departement_filter)

    # 🔹 Filtre par sous-département
    if sous_departement_filter:
        anciens = anciens.filter(adherant__sous_departement=sous_departement_filter)

    # 🔹 Pagination
    paginator = Paginator(anciens.order_by("-date_depart"), 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # 🔹 Récupérer les listes distinctes de départements et sous-départements
    departements = (
        AncienAdherent.objects.select_related('adherant')
        .filter(adherant__isnull=False)
        .values_list('adherant__departement', flat=True)
        .distinct()
    )
    sous_departements = (
        AncienAdherent.objects.select_related('adherant')
        .filter(adherant__isnull=False)
        .values_list('adherant__sous_departement', flat=True)
        .distinct()
    )

    return render(request, "decaissements/liste_anciens_adherents.html", {
        "page_obj": page_obj,
        "search_query": search_query,
        "departements": departements,
        "sous_departements": sous_departements,
        "departement_filter": departement_filter,
        "sous_departement_filter": sous_departement_filter,
    })


def export_anciens_adherents_pdf(request):
    search_query = request.GET.get("search", "")
    departement_filter = request.GET.get("departement", "")

    anciens = AncienAdherent.objects.select_related('adherant').all()

    if search_query:
        anciens = anciens.filter(
            Q(adherant__nom__icontains=search_query) |
            Q(adherant__prenom__icontains=search_query)
        )
    if departement_filter:
        anciens = anciens.filter(adherant__departement=departement_filter)

    # Réponse HTTP en PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="anciens_adherents.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Titre
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, height - 50, "Liste des Anciens Adhérents")

    y = height - 100
    p.setFont("Helvetica", 10)

    for ancien in anciens:
        adherant = ancien.adherant
        line = f"{adherant.nom} {adherant.prenom} | {adherant.departement} | {adherant.telephone} | {ancien.motif_depart} | {ancien.date_depart.strftime('%d/%m/%Y')}"
        p.drawString(50, y, line)
        y -= 20
        if y < 50:  # Nouvelle page
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 50

    p.save()
    return response

def export_anciens_adherents_excel(request):
    search_query = request.GET.get("search", "")
    departement_filter = request.GET.get("departement", "")

    anciens = AncienAdherent.objects.all()
    if search_query:
        anciens = anciens.filter(Q(nom__icontains=search_query) | Q(prenom__icontains=search_query))
    if departement_filter:
        anciens = anciens.filter(departement=departement_filter)

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anciens Adhérents"

    # Entêtes
    headers = ["Nom", "Prénom", "Email", "Téléphone", "Département", "Motif Départ", "Date Départ"]
    ws.append(headers)

    # Données
    for ancien in anciens:
        ws.append([
            ancien.nom,
            ancien.prenom,
            ancien.email,
            ancien.telephone,
            ancien.departement,
            ancien.motif_depart,
            ancien.date_depart.strftime("%d/%m/%Y"),
        ])

    # Réponse HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="anciens_adherents.xlsx"'
    wb.save(response)
    return response

def supprimer_ancien_adherent(request, id):
    ancien = get_object_or_404(AncienAdherent, id=id)

    if request.method == "POST":
        # ⚠️ Suppression définitive
        ancien.delete()
        messages.success(request, "Ancien adhérent supprimé définitivement.")

    return redirect('liste_anciens_adherents')



def ajouter_information(request):
    if request.method == 'POST':
        form = InformationForm(request.POST, request.FILES)  # important d'ajouter request.FILES
        if form.is_valid():
            form.save()
            return redirect('dashboard')  # ou la page où tu veux rediriger
    else:
        form = InformationForm()
    return render(request, 'app_social/ajouter_information.html', {'form': form})

def can_add_info(user):
    if user.is_superuser:
        return True
    return user.groups.filter(name="President").exists()

def archiver_adherent(request, id):
    adherant = get_object_or_404(Adherant, id=id)

    AncienAdherent.objects.create(
        adherant=adherant,
        date_depart=date.today(),
        motif_depart="Fin d'adhésion"
    )

    return redirect('liste_adherants')


# ============================================================
# GESTION DES COMPTES ADHÉRENTS — À ajouter dans views.py
# ============================================================
# ---- Liste des comptes ----

@login_required
@user_passes_test(lambda u: u.is_superuser)
def gestion_comptes(request):
    adherants = Adherant.objects.all().order_by('nom', 'prenom')
    groupes = Group.objects.all()

    adherants_data = []
    for a in adherants:
        compte = CompteAdherent.objects.filter(adherant=a).first()
        adherants_data.append({
            'adherant': a,
            'a_un_compte': compte is not None,
            'premiere_connexion': compte.premiere_connexion if compte else None,
            'username': a.user.username if a.user else None,
            'role': a.user.groups.first().name if a.user and a.user.groups.exists() else None,
        })

    context = {
        'adherants_data': adherants_data,
        'total': adherants.count(),
        'avec_compte': CompteAdherent.objects.count(),
        'sans_compte': adherants.count() - CompteAdherent.objects.count(),
        'groupes': groupes,
    }
    return render(request, 'admin/gestion_comptes.html', context)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def assigner_role(request, adherant_id):
    adherant = get_object_or_404(Adherant, id=adherant_id)

    if not adherant.user:
        messages.error(request, f"❌ {adherant} n'a pas de compte utilisateur.")
        return redirect('gestion_comptes')

    if request.method == 'POST':
        role_id = request.POST.get('role_id')

        # Retirer tous les groupes existants
        adherant.user.groups.clear()

        if role_id:
            groupe = get_object_or_404(Group, id=role_id)
            adherant.user.groups.add(groupe)
            messages.success(request, f"✅ Rôle '{groupe.name}' assigné à {adherant}.")
        else:
            messages.info(request, f"ℹ️ Rôle retiré pour {adherant}.")

    return redirect('gestion_comptes')    

# ---- Créer un compte pour un adhérent ----
@login_required
@user_passes_test(lambda u: u.is_staff)
def creer_compte_adherent(request, adherant_id):
    adherant = get_object_or_404(Adherant, id=adherant_id)

    # Vérifier qu'il n'a pas déjà un compte
    if hasattr(adherant, 'compte'):
        messages.warning(request, f"{adherant} a déjà un compte.")
        return redirect('gestion_comptes')

    # Vérifier que le téléphone n'est pas déjà utilisé comme username
    if User.objects.filter(username=adherant.telephone).exists():
        messages.error(request, f"Le numéro {adherant.telephone} est déjà utilisé par un autre compte.")
        return redirect('gestion_comptes')

    try:
        # Créer le User Django avec le téléphone comme username
        user = User.objects.create_user(
        username=adherant.telephone,
        password='passer123',
        first_name=adherant.prenom,
        last_name=adherant.nom,
        is_staff=False,      # ← ajoute ceci
        is_superuser=False,  # ← ajoute ceci
    )

        # Lier le user à l'adhérent
        adherant.user = user
        adherant.save()

        # Créer le CompteAdherent
        CompteAdherent.objects.create(
            adherant=adherant,
            premiere_connexion=True
        )

        messages.success(request, f"Compte créé pour {adherant}. Login : {adherant.telephone} / Mot de passe : passer123")

    except IntegrityError:
        messages.error(request, "Erreur lors de la création du compte.")

    return redirect('gestion_comptes')


# ---- Réinitialiser le mot de passe ----
@login_required
@user_passes_test(lambda u: u.is_staff)
def reinitialiser_mot_de_passe(request, adherant_id):
    adherant = get_object_or_404(Adherant, id=adherant_id)

    if not adherant.user:
        messages.error(request, f"{adherant} n'a pas de compte.")
        return redirect('gestion_comptes')

    # Remettre le mot de passe par défaut et forcer le changement
    adherant.user.set_password('passer123')
    adherant.user.save()

    # Remettre premiere_connexion à True
    compte = CompteAdherent.objects.filter(adherant=adherant).first()
    if compte:
        compte.premiere_connexion = True
        compte.save()

    messages.success(request, f"Mot de passe réinitialisé pour {adherant}. Il devra changer son mot de passe à la prochaine connexion.")
    return redirect('gestion_comptes')


# ---- Désactiver un compte ----
@login_required
@user_passes_test(lambda u: u.is_staff)
def desactiver_compte(request, adherant_id):
    adherant = get_object_or_404(Adherant, id=adherant_id)

    if not adherant.user:
        messages.error(request, f"{adherant} n'a pas de compte.")
        return redirect('gestion_comptes')

    adherant.user.is_active = not adherant.user.is_active
    adherant.user.save()

    statut = "activé" if adherant.user.is_active else "désactivé"
    messages.success(request, f"Compte {statut} pour {adherant}.")
    return redirect('gestion_comptes')


# ============================================================
# VUES PORTAIL ADHÉRENT — À ajouter dans views.py
# ============================================================

from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from .models import CompteAdherent
from django.contrib.auth import logout as auth_logout

def logout_view(request):
    auth_logout(request)
    return redirect('index')

def login_admin(request):
    if request.user.is_authenticated:
        est_bureau = request.user.groups.filter(name__in=ROLES_BUREAU).exists()
        if request.user.is_staff or est_bureau:
            return redirect('dashboard')

    erreur = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()

        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            est_bureau = user.groups.filter(name__in=ROLES_BUREAU).exists()
            if user.is_staff or est_bureau:
                login(request, user)
                return redirect('dashboard')
            else:
                erreur = "Accès non autorisé."
        else:
            erreur = "Identifiants incorrects."

    return render(request, 'admin/login_admin.html', {'erreur': erreur})

# ============================================================
# FONCTION UTILITAIRE — récupérer l'IP du client
# ============================================================
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


# ============================================================
# VUE 1 — LOGIN ADHÉRENT
# ============================================================
def login_adherent(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('dashboard')
        return redirect('portail_adherent')

    erreur = None
    if request.method == 'POST':
        telephone = request.POST.get('telephone', '').strip()
        password = request.POST.get('password', '').strip()

        user = authenticate(request, username=telephone, password=password)

        if user is not None:
            if not user.is_active:
                erreur = "Votre compte est désactivé. Contactez l'administrateur."
            else:
                login(request, user)

                # ✅ LOG DE CONNEXION
                try:
                    adherant = Adherant.objects.get(user=user)
                    LogConnexionAdherent.objects.create(
                        adherant=adherant,
                        action="Connexion au portail adhérent",
                        ip=get_client_ip(request)
                    )
                except Adherant.DoesNotExist:
                    pass

                # Vérifier si c'est la première connexion
                try:
                    compte = CompteAdherent.objects.get(adherant__user=user)
                    if compte.premiere_connexion:
                        return redirect('changer_mot_de_passe_adherent')
                except CompteAdherent.DoesNotExist:
                    pass

                next_url = request.GET.get('next', 'portail_adherent')
                return redirect(next_url)
        else:
            erreur = "Numéro de téléphone ou mot de passe incorrect."

    return render(request, 'adherent/login_adherent.html', {'erreur': erreur})


# ============================================================
# VUE 2 — LOGOUT ADHÉRENT
# ============================================================
def logout_adherent(request):
    # ✅ LOG DE DÉCONNEXION
    if request.user.is_authenticated:
        try:
            adherant = Adherant.objects.get(user=request.user)
            LogConnexionAdherent.objects.create(
                adherant=adherant,
                action="Déconnexion du portail adhérent",
                ip=get_client_ip(request)
            )
        except Adherant.DoesNotExist:
            pass

    logout(request)
    return redirect('index')


# ============================================================
# VUE 3 — CHANGEMENT DE MOT DE PASSE
# ============================================================
@login_required(login_url='login_adherent')
def changer_mot_de_passe_adherent(request):
    try:
        adherant = Adherant.objects.get(user=request.user)
        compte = CompteAdherent.objects.get(adherant=adherant)
    except (Adherant.DoesNotExist, CompteAdherent.DoesNotExist):
        return redirect('login_adherent')

    erreur = None

    if request.method == 'POST':
        telephone_saisi = request.POST.get('telephone', '').strip()
        nouveau_mdp = request.POST.get('nouveau_mdp', '')
        confirmer_mdp = request.POST.get('confirmer_mdp', '')

        if telephone_saisi != adherant.telephone:
            erreur = "Le numéro de téléphone ne correspond pas à celui enregistré."
        elif len(nouveau_mdp) < 6:
            erreur = "Le mot de passe doit contenir au moins 6 caractères."
        elif nouveau_mdp != confirmer_mdp:
            erreur = "Les mots de passe ne correspondent pas."
        elif nouveau_mdp == 'passer123':
            erreur = "Vous devez choisir un mot de passe différent du mot de passe par défaut."
        else:
            request.user.set_password(nouveau_mdp)
            request.user.save()

            compte.premiere_connexion = False
            compte.save()

            update_session_auth_hash(request, request.user)

            # ✅ LOG DE CHANGEMENT DE MOT DE PASSE
            LogConnexionAdherent.objects.create(
                adherant=adherant,
                action="Changement de mot de passe effectué",
                ip=get_client_ip(request)
            )

            messages.success(request, "Mot de passe changé avec succès !")
            return redirect('portail_adherent')

    return render(request, 'adherent/changer_mot_de_passe.html', {
        'erreur': erreur,
        'premiere_connexion': compte.premiere_connexion,
    })


# ============================================================
# VUE 4 — PORTAIL ADHÉRENT
# ============================================================
@login_required(login_url='login_adherent')
def portail_adherent(request):
    est_bureau = request.user.groups.filter(name__in=ROLES_BUREAU).exists()
    if request.user.is_staff:
        return redirect('dashboard')
    # ✅ Le bureau passe, les adhérents aussi
    try:
        adherant = Adherant.objects.get(user=request.user)
        compte = CompteAdherent.objects.get(adherant=adherant)
    except (Adherant.DoesNotExist, CompteAdherent.DoesNotExist):
        messages.error(request, "Accès réservé aux adhérents.")
        return redirect('login_adherent')

    if compte.premiere_connexion:
        return redirect('changer_mot_de_passe_adherent')

    LogConnexionAdherent.objects.create(
        adherant=adherant,
        action="Visite du portail adhérent",
        ip=get_client_ip(request)
    )

    return render(request, 'adherent/portail_adherent.html', {
        'adherant': adherant,
    })


# ============================================================
# VUE 5 — DEMANDE DE PRÊT PROTÉGÉE
# ============================================================
@login_required(login_url='login_adherent')
def demande_pret_adherent(request):

    try:
        adherant = Adherant.objects.get(user=request.user)
        compte = CompteAdherent.objects.get(adherant=adherant)

    except (Adherant.DoesNotExist, CompteAdherent.DoesNotExist):
        messages.error(request, "Accès réservé aux adhérents.")
        return redirect('login_adherent')

    # =========================================
    # FORCER CHANGEMENT MOT DE PASSE
    # =========================================
    if compte.premiere_connexion:
        return redirect('changer_mot_de_passe_adherent')

    # =========================================
    # MOIS COURANT
    # =========================================
    mois_fr = {
        1: "Janvier",
        2: "Février",
        3: "Mars",
        4: "Avril",
        5: "Mai",
        6: "Juin",
        7: "Juillet",
        8: "Août",
        9: "Septembre",
        10: "Octobre",
        11: "Novembre",
        12: "Décembre"
    }

    today = date.today()
    mois_courant = mois_fr[today.month]

    # =========================================
    # POST
    # =========================================
    if request.method == "POST":

        # =========================================
        # VERIFICATION COTISATION
        # =========================================
        cotisation_a_jour = CotisationMensuelle.objects.filter(
            adherant=adherant,
            mois=mois_courant,
            annee=today.year
        ).exists()

        # ❌ REFUS SI PAS A JOUR
        if not cotisation_a_jour:

            LogConnexionAdherent.objects.create(
                adherant=adherant,
                action="Tentative de demande de prêt refusée — cotisation non à jour",
                ip=get_client_ip(request)
            )

            messages.error(
                request,
                "❌ Demande rejetée. "
                "Vous devez régulariser votre cotisation du mois en cours avant de demander un prêt."
            )

            return redirect('demande_pret_adherent')

        # =========================================
        # FORMULAIRE
        # =========================================
        form = DemandePretForm(request.POST)

        if form.is_valid():

            pret = form.save(commit=False)

            # =========================================
            # INFOS ADHERENT
            # =========================================
            pret.adherant = adherant
            pret.nom = adherant.nom
            pret.prenom = adherant.prenom
            pret.telephone = adherant.telephone
            pret.departement = adherant.departement
            pret.adresse_rue = adherant.adresse_rue

            # =========================================
            # STATUT
            # =========================================
            pret.statut = "En attente"

            pret.save()

            # =========================================
            # LOG SUCCES
            # =========================================
            LogConnexionAdherent.objects.create(
                adherant=adherant,
                action=f"Demande de prêt soumise — Montant : {int(pret.montant):,} FCFA".replace(',', ' '),
                ip=get_client_ip(request)
            )

            messages.success(
                request,
                "✅ Votre demande de prêt a été envoyée avec succès."
            )

            return redirect('etat_pret_adherent')

        else:

            # =========================================
            # LOG ECHEC
            # =========================================
            LogConnexionAdherent.objects.create(
                adherant=adherant,
                action="Tentative de demande de prêt — formulaire invalide",
                ip=get_client_ip(request)
            )

            messages.error(
                request,
                "❌ Formulaire invalide. Vérifiez les champs."
            )

    else:

        # =========================================
        # LOG VISITE
        # =========================================
        LogConnexionAdherent.objects.create(
            adherant=adherant,
            action="Consultation du formulaire de demande de prêt",
            ip=get_client_ip(request)
        )

        # =========================================
        # PRE-REMPLISSAGE
        # =========================================
        form = DemandePretForm(initial={
            'nom': adherant.nom,
            'prenom': adherant.prenom,
            'telephone': adherant.telephone,
            'adresse_rue': adherant.adresse_rue,
            'departement': adherant.departement,
        })

    # =========================================
    # RENDER
    # =========================================
    return render(request, 'adherent/demande_pret_adherent.html', {
        'form': form,
        'adherant': adherant,
        'mois_courant': mois_courant,
    })

# ============================================================
# VUE 6 — CONSULTATION ÉTAT DU PRÊT
# ============================================================
@login_required(login_url='login_adherent')
def etat_pret_adherent(request):
    try:
        adherant = Adherant.objects.get(user=request.user)
    except Adherant.DoesNotExist:
        messages.error(request, "Accès réservé aux adhérents.")
        return redirect('login_adherent')

    # ✅ CompteAdherent optionnel pour les membres du bureau
    compte = CompteAdherent.objects.filter(adherant=adherant).first()
    if compte and compte.premiere_connexion:
        return redirect('changer_mot_de_passe_adherent')

    demandes_pret = DemandePret.objects.filter(adherant=adherant).order_by('-date_demande')
    cotisations = CotisationMensuelle.objects.filter(adherant=adherant).order_by('-annee', '-mois')

    mois_mapping = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril", 5: "Mai", 6: "Juin",
        7: "Juillet", 8: "Aout", 9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Decembre"
    }
    mois_actuel = mois_mapping[date.today().month]
    annee_actuelle = date.today().year

    cotisation_mois = cotisations.filter(
        mois=mois_actuel,
        annee=annee_actuelle
    ).first()

    a_jour = cotisation_mois is not None

    LogConnexionAdherent.objects.create(
        adherant=adherant,
        action="Consultation de l'état des cotisations et prêts",
        ip=get_client_ip(request)
    )

    return render(request, 'adherent/etat_pret_adherent.html', {
        'adherant': adherant,
        'demandes_pret': demandes_pret,
        'cotisations': cotisations,
        'a_jour': a_jour,
        'mois_actuel': mois_actuel,
        'annee_actuelle': annee_actuelle,
    })

# ============================================================
# VUE 7 — LOGS ADHÉRENTS (ADMIN)
# ============================================================
@login_required(login_url='login_admin')
def logs_adherents(request):
    est_autorise = request.user.is_staff or request.user.is_superuser
    if not est_autorise:
        messages.error(request, "⛔ Accès réservé à l'administrateur.")
        return redirect('dashboard')

    adherant_id = request.GET.get('adherant_id', '')
    logs = LogConnexionAdherent.objects.select_related('adherant').all()

    if adherant_id:
        logs = logs.filter(adherant__id=adherant_id)

    adherants = Adherant.objects.filter(compte__isnull=False).order_by('nom')

    context = {
        'logs': logs[:100],
        'adherants': adherants,
        'adherant_id': adherant_id,
    }
    return render(request, 'admin/logs_adherents.html', context)

def reinitialiser_mot_de_passe_adherent(request):
    erreur = None
    succes = None

    if request.method == 'POST':
        telephone = request.POST.get('telephone', '').strip()
        nouveau_mdp = request.POST.get('nouveau_mdp', '').strip()
        confirmer_mdp = request.POST.get('confirmer_mdp', '').strip()

        # Vérifier que l'adhérent existe
        try:
            adherant = Adherant.objects.get(telephone=telephone)
        except Adherant.DoesNotExist:
            erreur = "❌ Aucun compte trouvé avec ce numéro de téléphone."
            return render(request, 'adherent/reinitialiser_mdp.html', {'erreur': erreur})

        if not adherant.user:
            erreur = "❌ Ce numéro n'a pas de compte actif."
            return render(request, 'adherent/reinitialiser_mdp.html', {'erreur': erreur})

        if len(nouveau_mdp) < 6:
            erreur = "❌ Le mot de passe doit contenir au moins 6 caractères."
        elif nouveau_mdp != confirmer_mdp:
            erreur = "❌ Les mots de passe ne correspondent pas."
        elif nouveau_mdp == 'passer123':
            erreur = "❌ Choisissez un mot de passe différent du mot de passe par défaut."
        else:
            adherant.user.set_password(nouveau_mdp)
            adherant.user.save()

            # Remettre premiere_connexion à False
            compte = CompteAdherent.objects.filter(adherant=adherant).first()
            if compte:
                compte.premiere_connexion = False
                compte.save()

            # Log
            LogConnexionAdherent.objects.create(
                adherant=adherant,
                action="Réinitialisation du mot de passe depuis la page de connexion",
                ip=get_client_ip(request)
            )

            succes = "✅ Mot de passe modifié avec succès. Vous pouvez vous connecter."

    return render(request, 'adherent/reinitialiser_mdp.html', {
        'erreur': erreur,
        'succes': succes,
    })

